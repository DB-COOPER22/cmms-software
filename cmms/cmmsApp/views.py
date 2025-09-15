from django.shortcuts import render,redirect
from django.http import HttpResponse
from django.urls import reverse
from django.core.mail import send_mail, get_connection
from django.conf import settings
from pathlib import Path
from datetime import datetime
from threading import Thread
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
from .forms import ContactForm
from .utils_excel import append_submission_xlsx
from .utils_contact import normalize_phone_and_country, country_name_from_alpha2


EXCEL_DIR = os.path.join(settings.BASE_DIR, "data")
EXCEL_PATH = os.path.join(EXCEL_DIR, "carl_demo_requests.xlsx")

NAME_RE   = re.compile(r"^[A-Za-z\s'.-]{2,}$")
PHONE_RE  = re.compile(r"^\+?\d[\d\s\-()]{6,}$")

def _append_to_excel(row):
    os.makedirs(EXCEL_DIR, exist_ok=True)
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Requests"
        headers = ["Timestamp","Full Name","Company","Email","Country","Dial Code",
                   "Phone","Address","Message","Source IP"]
        ws.append(headers)
        for i in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(i)].width = 24
    ws.append(row)
    wb.save(EXCEL_PATH)

def request_demo_view(request: HttpRequest):
    if request.method == "POST":
        full_name = request.POST.get("full_name","").strip()
        company   = request.POST.get("company","").strip()
        email     = request.POST.get("email","").strip()
        phone     = request.POST.get("phone","").strip()
        country   = request.POST.get("country","").strip()  # "IN|+91"
        address   = request.POST.get("address","").strip()
        message   = request.POST.get("message","").strip()

        errors = {}
        if not NAME_RE.match(full_name):
            errors["full_name"] = "Enter a valid name (letters only)."

        if not company:
            errors["company"] = "Company is required."

        try:
            validate_email(email)
        except ValidationError:
            errors["email"] = "Enter a valid email address."

        if not PHONE_RE.match(phone):
            errors["phone"] = "Enter a valid phone number."

        if not country:
            errors["country"] = "Please select a country."

        # on error, re-render modal page if you use it standalone
        if errors:
            return render(request, "request_demo_popup.html", {"errors": errors})

        country_code, dial = (country.split("|", 1) + [""])[:2]

        _append_to_excel([
            datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
            full_name, company, email,
            country_code, dial, phone, address, message,
            request.META.get("REMOTE_ADDR",""),
        ])
        return redirect("contact_thanks")

    # GET fallback (if you open page directly)
    return render(request, "request_demo_popup.html")

def thanks_view(request):
    return render(request, "contact_thanks.html")



def home(request):
    return render(request, 'index.html')

def factory(request):
    return render(request, 'factory.html')

def healthcare(request):
    return render(request, 'healthcare.html')
def facility(request):
    return render(request, 'facility.html')
def city(request):
    return render(request, 'city.html')
def transport(request):
    return render(request, 'transport.html')
def contact(request):
    return render(request, 'contact.html')
def iot(request):
    return render(request, 'iot.html')
def eam(request):
    return render(request, 'eam.html')
def apm(request):
    return render(request, 'apm.html')
def mobility(request):
    return render(request, 'mobility.html')
def plans(request):
    return render(request, 'plans.html')

def about(request):
    return render(request, 'about.html')
def workorder(request):
    return render(request, "workorder.html")
def compliance(request):
    return render(request, "compliance.html")
def scada(request):
    return render(request, "scada.html")
def gis(request):
    return render(request, "gis.html")
def erpsync(request):
    return render(request, "erpsync.html")


def _send_contact_email_async(subject, body):
    """Send email in background; print any error to console."""
    try:
        conn = get_connection(timeout=getattr(settings, "EMAIL_TIMEOUT", 15))
        send_mail(
            subject=subject,
            message=body,
            from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None)
                        or getattr(settings, "EMAIL_HOST_USER", None),
            recipient_list=getattr(settings, "CONTACT_RECIPIENTS", None)
                        or [getattr(settings, "CONTACT_INBOX", None)
                        or getattr(settings, "EMAIL_HOST_USER", None)],
            fail_silently=False,
            connection=conn,
            auth_user=getattr(settings, "EMAIL_HOST_USER", None),
            auth_password=getattr(settings, "EMAIL_HOST_PASSWORD", None),
        )
    except Exception as e:
        print("EMAIL ERROR:", repr(e))

def contact_section(request):
    form = ContactForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        cd = form.cleaned_data

        # Normalize phone + resolve country
        e164_phone, resolved_alpha2, resolved_country_name = normalize_phone_and_country(
            cd.get("phone", ""), cd.get("country", "")
        )

        # Append to Excel
        xlsx_path = Path(getattr(settings, "CONTACT_SUBMISSIONS_XLSX",
                          Path(settings.BASE_DIR) / "contact_submissions.xlsx"))
        append_submission_xlsx(
            xlsx_path,
            [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                cd["first_name"],
                cd.get("last_name", ""),
                cd.get("company", ""),
                cd["email"],
                resolved_alpha2,
                resolved_country_name,
                e164_phone or cd.get("phone", ""),
                cd.get("message", ""),
            ],
        )

        # Email body
        body = "\n".join([
            "New contact submission for Carl  Software:",
            f"Name: {cd['first_name']} {cd.get('last_name','')}".strip(),
            f"Company: {cd.get('company','')}",
            f"Email: {cd['email']}",
            f"Country: {resolved_country_name or country_name_from_alpha2(resolved_alpha2) or cd.get('country','')}",
            f"Phone: {e164_phone or cd.get('phone','')}",
            "",
            "Message:",
            cd.get("message", "")
        ])

        # Fire-and-forget email
        Thread(target=_send_contact_email_async,
               args=("New website contact submission for Carl Software", body),
               daemon=True).start()

        # Redirect to Thanks page
        return redirect(reverse("cmmsApp:contact_thanks"))

    return render(request, "contact_section.html", {"form": form, "sent": request.GET.get("sent")})

def contact_thanks(request):
    return render(request, "contact_thanks.html", {})

